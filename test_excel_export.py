"""
Test script for Excel export functionality
Tests the license plate Excel export feature (Agent 2)
"""

import os
import sys
from util import write_license_plates_to_excel

# Create sample results data similar to what main.py generates
def create_sample_results():
    """Create sample results data for testing"""
    results = {
        0: {
            1: {
                'car': {'bbox': [100, 100, 200, 200]},
                'license_plate': {
                    'bbox': [120, 120, 180, 150],
                    'text': 'ABC1234',
                    'bbox_score': 0.95,
                    'text_score': 0.92
                }
            }
        },
        5: {
            1: {
                'car': {'bbox': [105, 105, 205, 205]},
                'license_plate': {
                    'bbox': [125, 125, 185, 155],
                    'text': 'ABC1234',
                    'bbox_score': 0.94,
                    'text_score': 0.91
                }
            }
        },
        10: {
            2: {
                'car': {'bbox': [300, 300, 400, 400]},
                'license_plate': {
                    'bbox': [320, 320, 380, 350],
                    'text': 'XYZ5678',
                    'bbox_score': 0.96,
                    'text_score': 0.93
                }
            }
        },
        15: {
            1: {
                'car': {'bbox': [110, 110, 210, 210]},
                'license_plate': {
                    'bbox': [130, 130, 190, 160],
                    'text': 'ABC1234',
                    'bbox_score': 0.93,
                    'text_score': 0.90
                }
            },
            2: {
                'car': {'bbox': [305, 305, 405, 405]},
                'license_plate': {
                    'bbox': [325, 325, 385, 355],
                    'text': 'XYZ5678',
                    'bbox_score': 0.95,
                    'text_score': 0.92
                }
            }
        }
    }
    return results

def test_excel_export():
    """Test the Excel export functionality"""
    print("Testing Excel export functionality...")
    
    # Create sample results
    results = create_sample_results()
    
    # Test file path
    test_output_path = './test_license_plates.xlsx'
    
    try:
        # Call the export function
        write_license_plates_to_excel(results, test_output_path)
        
        # Check if file was created
        if os.path.exists(test_output_path):
            file_size = os.path.getsize(test_output_path)
            print(f"✓ Excel file created successfully: {test_output_path}")
            print(f"  File size: {file_size} bytes")
            
            # Verify file can be opened with openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(test_output_path)
            ws = wb.active
            
            # Check headers
            headers = [cell.value for cell in ws[1]]
            expected_headers = ['License Plate Number', 'First Detected Frame', 
                              'Last Detected Frame', 'Total Detections', 'Average Confidence']
            
            if headers == expected_headers:
                print("✓ Headers are correct")
            else:
                print(f"✗ Headers mismatch. Expected: {expected_headers}, Got: {headers}")
                return False
            
            # Check data rows (should have 2 unique plates: ABC1234 and XYZ5678)
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            print(f"✓ Found {len(data_rows)} unique license plates")
            
            for row in data_rows:
                print(f"  - {row[0]}: frames {row[1]}-{row[2]}, {row[3]} detections, avg confidence {row[4]}")
            
            # Clean up test file
            os.remove(test_output_path)
            print(f"✓ Test file cleaned up")
            
            print("\n✓ All tests passed!")
            return True
            
        else:
            print(f"✗ Excel file was not created")
            return False
            
    except Exception as e:
        print(f"✗ Error during testing: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_excel_export()
    sys.exit(0 if success else 1)
