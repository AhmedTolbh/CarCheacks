"""
Simple test script for Excel export functionality
Tests only the Excel writing part without importing easyocr
"""

import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def write_license_plates_to_excel_test(results, output_path):
    """
    Test version of Excel export function (copied to avoid easyocr import)
    """
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "License Plates"
    
    # Set up header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Write headers
    headers = ['License Plate Number', 'First Detected Frame', 'Last Detected Frame', 
               'Total Detections', 'Average Confidence']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Track unique license plates and their data
    license_plate_data = {}
    
    # Process results to extract unique license plates
    for frame_nmr in results.keys():
        for car_id in results[frame_nmr].keys():
            if 'car' in results[frame_nmr][car_id].keys() and \
               'license_plate' in results[frame_nmr][car_id].keys() and \
               'text' in results[frame_nmr][car_id]['license_plate'].keys():
                
                plate_text = results[frame_nmr][car_id]['license_plate']['text']
                confidence = results[frame_nmr][car_id]['license_plate']['text_score']
                
                if plate_text not in license_plate_data:
                    license_plate_data[plate_text] = {
                        'first_frame': frame_nmr,
                        'last_frame': frame_nmr,
                        'detections': 1,
                        'confidences': [confidence]
                    }
                else:
                    license_plate_data[plate_text]['last_frame'] = frame_nmr
                    license_plate_data[plate_text]['detections'] += 1
                    license_plate_data[plate_text]['confidences'].append(confidence)
    
    # Write data rows
    row_num = 2
    for plate_text, data in sorted(license_plate_data.items()):
        avg_confidence = sum(data['confidences']) / len(data['confidences'])
        
        ws.cell(row=row_num, column=1, value=plate_text)
        ws.cell(row=row_num, column=2, value=data['first_frame'])
        ws.cell(row=row_num, column=3, value=data['last_frame'])
        ws.cell(row=row_num, column=4, value=data['detections'])
        ws.cell(row=row_num, column=5, value=f"{avg_confidence:.2f}")
        
        row_num += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_path)
    print(f"Excel file saved to {output_path} with {len(license_plate_data)} unique license plates")

# Create sample results data similar to what main.py generates
def create_sample_results():
    """Create sample results data for testing"""
    results = {
        0: {
            1: {
                'car': {'bbox': [100, 100, 200, 200]},
                'license_plate': {
                    'bbox': [120, 120, 180, 150],
                    'text': 'ABC1234',
                    'bbox_score': 0.95,
                    'text_score': 0.92
                }
            }
        },
        5: {
            1: {
                'car': {'bbox': [105, 105, 205, 205]},
                'license_plate': {
                    'bbox': [125, 125, 185, 155],
                    'text': 'ABC1234',
                    'bbox_score': 0.94,
                    'text_score': 0.91
                }
            }
        },
        10: {
            2: {
                'car': {'bbox': [300, 300, 400, 400]},
                'license_plate': {
                    'bbox': [320, 320, 380, 350],
                    'text': 'XYZ5678',
                    'bbox_score': 0.96,
                    'text_score': 0.93
                }
            }
        },
        15: {
            1: {
                'car': {'bbox': [110, 110, 210, 210]},
                'license_plate': {
                    'bbox': [130, 130, 190, 160],
                    'text': 'ABC1234',
                    'bbox_score': 0.93,
                    'text_score': 0.90
                }
            },
            2: {
                'car': {'bbox': [305, 305, 405, 405]},
                'license_plate': {
                    'bbox': [325, 325, 385, 355],
                    'text': 'XYZ5678',
                    'bbox_score': 0.95,
                    'text_score': 0.92
                }
            }
        },
        20: {
            3: {
                'car': {'bbox': [500, 500, 600, 600]},
                'license_plate': {
                    'bbox': [520, 520, 580, 550],
                    'text': 'DEF9876',
                    'bbox_score': 0.97,
                    'text_score': 0.94
                }
            }
        }
    }
    return results

def test_excel_export():
    """Test the Excel export functionality"""
    print("Testing Excel export functionality...")
    
    # Create sample results
    results = create_sample_results()
    
    # Test file path
    test_output_path = './test_license_plates.xlsx'
    
    try:
        # Call the export function
        write_license_plates_to_excel_test(results, test_output_path)
        
        # Check if file was created
        if os.path.exists(test_output_path):
            file_size = os.path.getsize(test_output_path)
            print(f"✓ Excel file created successfully: {test_output_path}")
            print(f"  File size: {file_size} bytes")
            
            # Verify file can be opened with openpyxl
            wb = load_workbook(test_output_path)
            ws = wb.active
            
            # Check headers
            headers = [cell.value for cell in ws[1]]
            expected_headers = ['License Plate Number', 'First Detected Frame', 
                              'Last Detected Frame', 'Total Detections', 'Average Confidence']
            
            if headers == expected_headers:
                print("✓ Headers are correct")
            else:
                print(f"✗ Headers mismatch. Expected: {expected_headers}, Got: {headers}")
                return False
            
            # Check data rows (should have 3 unique plates: ABC1234, DEF9876, and XYZ5678)
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            print(f"✓ Found {len(data_rows)} unique license plates")
            
            if len(data_rows) != 3:
                print(f"✗ Expected 3 unique plates, got {len(data_rows)}")
                return False
            
            for row in data_rows:
                print(f"  - {row[0]}: frames {row[1]}-{row[2]}, {row[3]} detections, avg confidence {row[4]}")
            
            # Verify specific data
            plate_dict = {row[0]: row for row in data_rows}
            
            # Check ABC1234 (should have 3 detections across frames 0, 5, 15)
            if 'ABC1234' in plate_dict:
                abc_row = plate_dict['ABC1234']
                if abc_row[1] == 0 and abc_row[2] == 15 and abc_row[3] == 3:
                    print("✓ ABC1234 data is correct")
                else:
                    print(f"✗ ABC1234 data incorrect: {abc_row}")
                    return False
            
            # Check XYZ5678 (should have 2 detections across frames 10, 15)
            if 'XYZ5678' in plate_dict:
                xyz_row = plate_dict['XYZ5678']
                if xyz_row[1] == 10 and xyz_row[2] == 15 and xyz_row[3] == 2:
                    print("✓ XYZ5678 data is correct")
                else:
                    print(f"✗ XYZ5678 data incorrect: {xyz_row}")
                    return False
            
            # Check DEF9876 (should have 1 detection at frame 20)
            if 'DEF9876' in plate_dict:
                def_row = plate_dict['DEF9876']
                if def_row[1] == 20 and def_row[2] == 20 and def_row[3] == 1:
                    print("✓ DEF9876 data is correct")
                else:
                    print(f"✗ DEF9876 data incorrect: {def_row}")
                    return False
            
            # Clean up test file
            os.remove(test_output_path)
            print(f"✓ Test file cleaned up")
            
            print("\n✓ All tests passed!")
            return True
            
        else:
            print(f"✗ Excel file was not created")
            return False
            
    except Exception as e:
        print(f"✗ Error during testing: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_excel_export()
    sys.exit(0 if success else 1)
